import pandas as pd
from openpyxl import load_workbook

# Excel origen
file_path = "HRNUEVO.xlsm"

# Excel destino
file_path2 = "DATOS HR y CONTRATOS.xlsx"
wb = load_workbook(file_path2)

# Metemos los datos en la hoja "HR"
ws = wb["HR"]  

# fila en la que se introducen los datos
n=int(input("En que fila se deben introducir los datos")) 

# Dataframes
modelo_df = pd.read_excel(file_path, sheet_name="MODELO", header=None)
tabla_balance_df = pd.read_excel(file_path, sheet_name="TABLA BALANCE", header=None)
hr_df = pd.read_excel(file_path, sheet_name="6. HR", header=None)

# Celdas O1 y G5
o1_value = modelo_df.iloc[0, 14]
g5_value = modelo_df.iloc[4, 6]

# Extraccion de celdas
celdaE2 = modelo_df.iloc[1, 4]
celdaG4 = modelo_df.iloc[3, 6]
celdaG5 = modelo_df.iloc[4, 6]
celdaG55 = modelo_df.iloc[54, 6]
celdaG58 = modelo_df.iloc[57, 6]
celdaG40 = modelo_df.iloc[39, 6]
celdaG41 = modelo_df.iloc[40, 6]
celdaG42 = modelo_df.iloc[41, 6]
celdaG43 = modelo_df.iloc[42, 6]
celdaG44 = modelo_df.iloc[43, 6]
celdaG45 = modelo_df.iloc[44, 6]
celdaG46 = modelo_df.iloc[45, 6]
celdaG47 = modelo_df.iloc[46, 6]
celdaG48 = modelo_df.iloc[47, 6]
celdaG33 = modelo_df.iloc[32, 6]
celdaG34 = modelo_df.iloc[33, 6]
celdaG35 = modelo_df.iloc[34, 6]
celdaG36 = modelo_df.iloc[35, 6]
celdaG37 = modelo_df.iloc[36, 6]
celdaG38 = modelo_df.iloc[37, 6]
celdaO15 = modelo_df.iloc[14, 14]
celdaO16 = modelo_df.iloc[15, 14]
celdaO20 = modelo_df.iloc[19, 14]
celdaO21 = modelo_df.iloc[21, 14]

celdaB4TABLABALANCE = tabla_balance_df.iloc[3, 1]
celdaB5TABLABALANCE = tabla_balance_df.iloc[4, 1]
celdaB6TABLABALANCE = tabla_balance_df.iloc[5, 1]
celdaB7TABLABALANCE = tabla_balance_df.iloc[6, 1]
celdaB8TABLABALANCE = tabla_balance_df.iloc[7, 1]
celdaB9TABLABALANCE = tabla_balance_df.iloc[8, 1]
celdaB10TABLABALANCE = tabla_balance_df.iloc[9, 1]
celdaB11TABLABALANCE = tabla_balance_df.iloc[10, 1]
celdaB12TABLABALANCE = tabla_balance_df.iloc[11, 1]
celdaB13TABLABALANCE = tabla_balance_df.iloc[12, 1]
celdaB14TABLABALANCE = tabla_balance_df.iloc[13, 1]
celdaB15TABLABALANCE = tabla_balance_df.iloc[14, 1]
celdaB16TABLABALANCE = tabla_balance_df.iloc[15, 1]
celdaC4TABLABALANCE = tabla_balance_df.iloc[3, 2]
celdaC5TABLABALANCE = tabla_balance_df.iloc[4, 2]
celdaC6TABLABALANCE = tabla_balance_df.iloc[5, 2]
celdaC7TABLABALANCE = tabla_balance_df.iloc[6, 2]
celdaC8TABLABALANCE = tabla_balance_df.iloc[7, 2]
celdaC9TABLABALANCE = tabla_balance_df.iloc[8, 2]
celdaC10TABLABALANCE = tabla_balance_df.iloc[9, 2]
celdaC11TABLABALANCE = tabla_balance_df.iloc[10, 2]
celdaC12TABLABALANCE = tabla_balance_df.iloc[11, 2]
celdaC13TABLABALANCE = tabla_balance_df.iloc[12, 2]
celdaC14TABLABALANCE = tabla_balance_df.iloc[13, 2]
celdaC15TABLABALANCE = tabla_balance_df.iloc[14, 2]
celdaC16TABLABALANCE = tabla_balance_df.iloc[15, 2]
celdaD4TABLABALANCE = tabla_balance_df.iloc[3, 3]
celdaD5TABLABALANCE = tabla_balance_df.iloc[4, 3]
celdaD6TABLABALANCE = tabla_balance_df.iloc[5, 3]
celdaD7TABLABALANCE = tabla_balance_df.iloc[6, 3]
celdaD8TABLABALANCE = tabla_balance_df.iloc[7, 3]
celdaD9TABLABALANCE = tabla_balance_df.iloc[8, 3]
celdaD10TABLABALANCE = tabla_balance_df.iloc[9, 3]
celdaD11TABLABALANCE = tabla_balance_df.iloc[10, 3]
celdaD12TABLABALANCE = tabla_balance_df.iloc[11, 3]
celdaD13TABLABALANCE = tabla_balance_df.iloc[12, 3]
celdaD14TABLABALANCE = tabla_balance_df.iloc[13, 3]
celdaD15TABLABALANCE = tabla_balance_df.iloc[14, 3]
celdaD16TABLABALANCE = tabla_balance_df.iloc[15, 3]
celdaE4TABLABALANCE = tabla_balance_df.iloc[3, 4]
celdaE5TABLABALANCE = tabla_balance_df.iloc[4, 4]
celdaE6TABLABALANCE = tabla_balance_df.iloc[5, 4]
celdaE7TABLABALANCE = tabla_balance_df.iloc[6, 4]
celdaE8TABLABALANCE = tabla_balance_df.iloc[7, 4]
celdaE9TABLABALANCE = tabla_balance_df.iloc[8, 4]
celdaE10TABLABALANCE = tabla_balance_df.iloc[9, 4]
celdaE11TABLABALANCE = tabla_balance_df.iloc[10, 4]
celdaE12TABLABALANCE = tabla_balance_df.iloc[11, 4]
celdaE13TABLABALANCE = tabla_balance_df.iloc[12, 4]
celdaE14TABLABALANCE = tabla_balance_df.iloc[13, 4]
celdaE15TABLABALANCE = tabla_balance_df.iloc[14, 4]
celdaE16TABLABALANCE = tabla_balance_df.iloc[15, 4]

celdaO19 = modelo_df.iloc[18, 14]
celdaO21 = modelo_df.iloc[20, 14]
celdaO24 = modelo_df.iloc[23, 14]
celdaO42 = modelo_df.iloc[41, 14]
celdaO47 = modelo_df.iloc[46, 14]

celdaE3HR = hr_df.iloc[2, 4]
celdaC14HR = hr_df.iloc[13, 2]
celdaD8HR = hr_df.iloc[7, 3]
celdaE8HR = hr_df.iloc[7, 4]
celdaF8HR = hr_df.iloc[7, 5]
celdaG8HR = hr_df.iloc[7, 6]
celdaH8HR = hr_df.iloc[7, 7]
celdaI8HR = hr_df.iloc[7, 8]
celdaJ8HR = hr_df.iloc[7, 9]
celdaK8HR = hr_df.iloc[7, 10]
celdaL8HR = hr_df.iloc[7, 11]
celdaM8HR = hr_df.iloc[7, 12]
celdaN8HR = hr_df.iloc[7, 13]
celdaO8HR = hr_df.iloc[7, 14]
celdaP8HR = hr_df.iloc[7, 15]
celdaQ8HR = hr_df.iloc[7, 16]
celdaR8HR = hr_df.iloc[7, 17]
celdaD17HR = hr_df.iloc[16, 3]
celdaE17HR = hr_df.iloc[16, 4]
celdaF17HR = hr_df.iloc[16, 5]
celdaG17HR = hr_df.iloc[16, 6]
celdaH17HR = hr_df.iloc[16, 7]
celdaI17HR = hr_df.iloc[16, 8]
celdaJ17HR = hr_df.iloc[16, 9]
celdaK17HR = hr_df.iloc[16, 10]
celdaL17HR = hr_df.iloc[16, 11]
celdaM17HR = hr_df.iloc[16, 12]
celdaN17HR = hr_df.iloc[16, 13]
celdaO17HR = hr_df.iloc[16, 14]
celdaP17HR = hr_df.iloc[16, 15]
celdaQ17HR = hr_df.iloc[16, 16]
celdaR17HR = hr_df.iloc[16, 17]

celdaK3 = modelo_df.iloc[2, 10]
celdaK4 = modelo_df.iloc[3, 10]
celdaK5 = modelo_df.iloc[4, 10]
celdaK8 = modelo_df.iloc[7, 10]


# Condicion 1: O1 es "2024-16EC" o "2024-21EC"
if o1_value in ["2024-16EC", "2024-21EC"]:
    print("HOLA1")

    # Insertar los datos en el excel
    ws[f"D{n}"] = celdaE2
    ws[f"C{n}"] = celdaG5
    ws[f"T{n}"] = celdaG58
    ws[f"E{n}"] = celdaG40
    ws[f"I{n}"] = celdaG42
    ws[f"J{n}"] = celdaG43
    ws[f"H{n}"] = celdaG44
    ws[f"K{n}"] = celdaG45
    ws[f"L{n}"] = celdaG46
    ws[f"O{n}"] = celdaG33
    ws[f"P{n}"] = celdaG34
    ws[f"M{n}"] = celdaO15
    ws[f"N{n}"] = celdaO16

    ws[f"AZ{n}"] = celdaB4TABLABALANCE
    ws[f"BA{n}"] = celdaB5TABLABALANCE
    ws[f"BB{n}"] = celdaB6TABLABALANCE
    ws[f"BC{n}"] = celdaB7TABLABALANCE
    ws[f"BD{n}"] = celdaB8TABLABALANCE
    ws[f"BE{n}"] = celdaB9TABLABALANCE
    ws[f"BF{n}"] = celdaB10TABLABALANCE
    ws[f"BG{n}"] = celdaB11TABLABALANCE
    ws[f"BH{n}"] = celdaB12TABLABALANCE
    ws[f"BI{n}"] = celdaB13TABLABALANCE
    ws[f"BJ{n}"] = celdaB14TABLABALANCE
    ws[f"BK{n}"] = celdaB15TABLABALANCE
    ws[f"BL{n}"] = celdaC4TABLABALANCE
    ws[f"BM{n}"] = celdaC5TABLABALANCE
    ws[f"BN{n}"] = celdaC6TABLABALANCE
    ws[f"BO{n}"] = celdaC7TABLABALANCE
    ws[f"BP{n}"] = celdaC8TABLABALANCE
    ws[f"BQ{n}"] = celdaC9TABLABALANCE
    ws[f"BR{n}"] = celdaC10TABLABALANCE
    ws[f"BS{n}"] = celdaC11TABLABALANCE
    ws[f"BT{n}"] = celdaC12TABLABALANCE
    ws[f"BU{n}"] = celdaC13TABLABALANCE
    ws[f"BV{n}"] = celdaC14TABLABALANCE
    ws[f"BW{n}"] = celdaC15TABLABALANCE
    ws[f"BX{n}"] = celdaD4TABLABALANCE
    ws[f"BY{n}"] = celdaD5TABLABALANCE
    ws[f"BZ{n}"] = celdaD6TABLABALANCE
    ws[f"CA{n}"] = celdaD7TABLABALANCE
    ws[f"CB{n}"] = celdaD8TABLABALANCE
    ws[f"CC{n}"] = celdaD9TABLABALANCE
    ws[f"CD{n}"] = celdaD10TABLABALANCE
    ws[f"CE{n}"] = celdaD11TABLABALANCE
    ws[f"CF{n}"] = celdaD12TABLABALANCE
    ws[f"CG{n}"] = celdaD13TABLABALANCE
    ws[f"CH{n}"] = celdaD14TABLABALANCE
    ws[f"CI{n}"] = celdaD15TABLABALANCE
    ws[f"CJ{n}"] = celdaE4TABLABALANCE
    ws[f"CK{n}"] = celdaE5TABLABALANCE
    ws[f"CL{n}"] = celdaE6TABLABALANCE
    ws[f"CM{n}"] = celdaE7TABLABALANCE
    ws[f"CN{n}"] = celdaE8TABLABALANCE
    ws[f"CO{n}"] = celdaE9TABLABALANCE
    ws[f"CP{n}"] = celdaE10TABLABALANCE
    ws[f"CQ{n}"] = celdaE11TABLABALANCE
    ws[f"CR{n}"] = celdaE12TABLABALANCE
    ws[f"CS{n}"] = celdaE13TABLABALANCE
    ws[f"CT{n}"] = celdaE14TABLABALANCE
    ws[f"CU{n}"] = celdaE15TABLABALANCE
    
    # Condicion 2: G5 es "PPA"
    if g5_value == "PPA":

        ws[f"U{n}"] = celdaD8HR
        ws[f"V{n}"] = celdaE8HR
        ws[f"W{n}"] = celdaF8HR
        ws[f"X{n}"] = celdaG8HR
        ws[f"Y{n}"] = celdaH8HR
        ws[f"Z{n}"] = celdaI8HR
        ws[f"AA{n}"] = celdaJ8HR
        ws[f"AB{n}"] = celdaK8HR
        ws[f"AC{n}"] = celdaL8HR
        ws[f"AD{n}"] = celdaM8HR
        ws[f"AE{n}"] = celdaN8HR
        ws[f"AF{n}"] = celdaO8HR
        ws[f"AG{n}"] = celdaP8HR
        ws[f"AH{n}"] = celdaQ8HR
        ws[f"AI{n}"] = celdaR8HR
        ws[f"AK{n}"] = celdaD17HR
        ws[f"AL{n}"] = celdaE17HR
        ws[f"AM{n}"] = celdaF17HR
        ws[f"AN{n}"] = celdaG17HR
        ws[f"AO{n}"] = celdaH17HR
        ws[f"AP{n}"] = celdaI17HR
        ws[f"AQ{n}"] = celdaJ17HR
        ws[f"AR{n}"] = celdaK17HR
        ws[f"AS{n}"] = celdaL17HR
        ws[f"AT{n}"] = celdaM17HR
        ws[f"AU{n}"] = celdaN17HR
        ws[f"AV{n}"] = celdaO17HR
        ws[f"AW{n}"] = celdaP17HR
        ws[f"AX{n}"] = celdaQ17HR
        ws[f"AY{n}"] = celdaR17HR

    # Condicion 3: G5 es "VD"
    if g5_value == "VD":

        ws[f"G{n}"] = celdaK3
        ws[f"F{n}"] = celdaK8
        




# Condicion 1: O1 es "2023-43EC" o "2024-05EC"
if o1_value in ["2023-43EC", "2024-05EC"]:
    print("HOLA2")
    celdaE2 = modelo_df.iloc[1, 4]
    celdaG4 = modelo_df.iloc[3, 6]
    celdaG5 = modelo_df.iloc[4, 6]
    celdaG55 = modelo_df.iloc[54, 6]
    celdaG40 = modelo_df.iloc[39, 6]
    celdaG42 = modelo_df.iloc[41, 6]
    celdaG43 = modelo_df.iloc[42, 6]
    celdaG44 = modelo_df.iloc[43, 6]
    celdaG45 = modelo_df.iloc[44, 6]
    celdaG46 = modelo_df.iloc[45, 6]
    celdaG33 = modelo_df.iloc[32, 6]
    celdaG34 = modelo_df.iloc[33, 6]
    celdaG36 = modelo_df.iloc[35, 6]
    celdaO15 = modelo_df.iloc[14, 14]
    celdaO16 = modelo_df.iloc[15, 14]
    celdaB4TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB5TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB6TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB7TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB8TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB9TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB10TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB11TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB12TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB13TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB14TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB15TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB16TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB17TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaC4TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC5TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC6TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC7TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC8TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC9TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC10TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC11TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC12TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC13TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC14TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC15TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC16TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC17TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaD4TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD5TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD6TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD7TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD8TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD9TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD10TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD11TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD12TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD13TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD14TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD15TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD16TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD17TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaE4TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE5TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE6TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE7TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE8TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE9TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE10TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE11TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE12TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE13TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE14TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE15TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE16TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE17TABLABALANCE = tabla_balance_df.iloc[8, 4]

    # Insertar el valor en la celda J6 (columna J, fila 6)
    ws["J6"] = 1
    
    # Condicion 2: G5 es "PPA"
    if g5_value == "PPA":
        celdaO19 = modelo_df.iloc[18, 14]
        celdaO42 = modelo_df.iloc[41, 14]
        celdaE3HR = hr_df.iloc[2, 4]
        celdaC14HR = hr_df.iloc[13, 2]
        celdaD8HR = hr_df.iloc[7, 3]
        celdaE8HR = hr_df.iloc[7, 4]
        celdaF8HR = hr_df.iloc[7, 5]
        celdaG8HR = hr_df.iloc[7, 6]
        celdaH8HR = hr_df.iloc[7, 7]
        celdaI8HR = hr_df.iloc[7, 8]
        celdaJ8HR = hr_df.iloc[7, 9]
        celdaK8HR = hr_df.iloc[7, 10]
        celdaL8HR = hr_df.iloc[7, 11]
        celdaM8HR = hr_df.iloc[7, 12]
        celdaN8HR = hr_df.iloc[7, 13]
        celdaO8HR = hr_df.iloc[7, 14]
        celdaP8HR = hr_df.iloc[7, 15]
        celdaQ8HR = hr_df.iloc[7, 16]
        celdaR8HR = hr_df.iloc[7, 17]
        celdaD17HR = hr_df.iloc[16, 3]
        celdaE17HR = hr_df.iloc[16, 4]
        celdaF17HR = hr_df.iloc[16, 5]
        celdaG17HR = hr_df.iloc[16, 6]
        celdaH17HR = hr_df.iloc[16, 7]
        celdaI17HR = hr_df.iloc[16, 8]
        celdaJ17HR = hr_df.iloc[16, 9]
        celdaK17HR = hr_df.iloc[16, 10]
        celdaL17HR = hr_df.iloc[16, 11]
        celdaM17HR = hr_df.iloc[16, 12]
        celdaN17HR = hr_df.iloc[16, 13]
        celdaO17HR = hr_df.iloc[16, 14]
        celdaP17HR = hr_df.iloc[16, 15]
        celdaQ17HR = hr_df.iloc[16, 16]
        celdaR17HR = hr_df.iloc[16, 17]

        # Insertar el valor en la celda J6 (columna J, fila 6)
        ws["J6"] = 1

    # Condicion 3: G5 es "VD"
    if g5_value == "VD":
        celdaK3 = modelo_df.iloc[2, 10]
        celdaK8 = modelo_df.iloc[7, 10]

        # Insertar el valor en la celda J6 (columna J, fila 6)
        ws["J6"] = 1





# Condicion 1: O1 es "2023-27EC"
if o1_value in ["2023-27EC"]:
    print("HOLA3")
    celdaE2 = modelo_df.iloc[1, 4]
    celdaG4 = modelo_df.iloc[3, 6]
    celdaG5 = modelo_df.iloc[4, 6]
    celdaG55 = modelo_df.iloc[54, 6]
    celdaG40 = modelo_df.iloc[39, 6]
    celdaG42 = modelo_df.iloc[41, 6]
    celdaG43 = modelo_df.iloc[42, 6]
    celdaG44 = modelo_df.iloc[43, 6]
    celdaG45 = modelo_df.iloc[44, 6]
    celdaG46 = modelo_df.iloc[45, 6]
    celdaG33 = modelo_df.iloc[32, 6]
    celdaG34 = modelo_df.iloc[33, 6]
    celdaG36 = modelo_df.iloc[35, 6]
    celdaO15 = modelo_df.iloc[14, 14]
    celdaO16 = modelo_df.iloc[15, 14]
    celdaB4TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB5TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB6TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB7TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB8TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB9TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB10TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB11TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB12TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB13TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB14TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB15TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB16TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB17TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaC4TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC5TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC6TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC7TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC8TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC9TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC10TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC11TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC12TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC13TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC14TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC15TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC16TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC17TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaD4TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD5TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD6TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD7TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD8TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD9TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD10TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD11TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD12TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD13TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD14TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD15TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD16TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD17TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaE4TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE5TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE6TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE7TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE8TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE9TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE10TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE11TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE12TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE13TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE14TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE15TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE16TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE17TABLABALANCE = tabla_balance_df.iloc[8, 4]

    # Insertar el valor en la celda J6 (columna J, fila 6)
    ws["J6"] = 1
        
    # Condicion 2: G5 es "PPA"
    if g5_value == "PPA":
        celdaO19 = modelo_df.iloc[18, 14]
        celdaO42 = modelo_df.iloc[41, 14]
        celdaE3HR = hr_df.iloc[2, 4]
        celdaC14HR = hr_df.iloc[13, 2]
        celdaD8HR = hr_df.iloc[7, 3]
        celdaE8HR = hr_df.iloc[7, 4]
        celdaF8HR = hr_df.iloc[7, 5]
        celdaG8HR = hr_df.iloc[7, 6]
        celdaH8HR = hr_df.iloc[7, 7]
        celdaI8HR = hr_df.iloc[7, 8]
        celdaJ8HR = hr_df.iloc[7, 9]
        celdaK8HR = hr_df.iloc[7, 10]
        celdaL8HR = hr_df.iloc[7, 11]
        celdaM8HR = hr_df.iloc[7, 12]
        celdaN8HR = hr_df.iloc[7, 13]
        celdaO8HR = hr_df.iloc[7, 14]
        celdaP8HR = hr_df.iloc[7, 15]
        celdaQ8HR = hr_df.iloc[7, 16]
        celdaR8HR = hr_df.iloc[7, 17]
        celdaD17HR = hr_df.iloc[16, 3]
        celdaE17HR = hr_df.iloc[16, 4]
        celdaF17HR = hr_df.iloc[16, 5]
        celdaG17HR = hr_df.iloc[16, 6]
        celdaH17HR = hr_df.iloc[16, 7]
        celdaI17HR = hr_df.iloc[16, 8]
        celdaJ17HR = hr_df.iloc[16, 9]
        celdaK17HR = hr_df.iloc[16, 10]
        celdaL17HR = hr_df.iloc[16, 11]
        celdaM17HR = hr_df.iloc[16, 12]
        celdaN17HR = hr_df.iloc[16, 13]
        celdaO17HR = hr_df.iloc[16, 14]
        celdaP17HR = hr_df.iloc[16, 15]
        celdaQ17HR = hr_df.iloc[16, 16]
        celdaR17HR = hr_df.iloc[16, 17]

        # Insertar el valor en la celda J6 (columna J, fila 6)
        ws["J6"] = 1

    # Condicion 3: G5 es "VD"
    if g5_value == "VD":
        celdaK3 = modelo_df.iloc[2, 10]
        celdaK5 = modelo_df.iloc[4, 10]

        # Insertar el valor en la celda J6 (columna J, fila 6)
        ws["J6"] = 1





# Condicion 1: O1 es "2023-15EC"
if o1_value in ["2023-15EC"]:
    print("HOLA4")
    celdaE2 = modelo_df.iloc[1, 4]
    celdaG4 = modelo_df.iloc[3, 6]
    celdaG5 = modelo_df.iloc[4, 6]
    celdaG48 = modelo_df.iloc[47, 6]
    celdaG42 = modelo_df.iloc[41, 6]
    celdaG45 = modelo_df.iloc[44, 6]
    celdaG43 = modelo_df.iloc[42, 6]
    celdaG44 = modelo_df.iloc[43, 6]
    celdaG47 = modelo_df.iloc[46, 6]
    celdaG37 = modelo_df.iloc[36, 6]
    celdaG38 = modelo_df.iloc[37, 6]
    celdaG46 = modelo_df.iloc[45, 6]
    celdaO15 = modelo_df.iloc[14, 14]
    celdaO16 = modelo_df.iloc[15, 14]
    celdaB4TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB5TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB6TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB7TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB8TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB9TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB10TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB11TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB12TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB13TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB14TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB15TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB16TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB17TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaC4TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC5TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC6TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC7TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC8TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC9TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC10TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC11TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC12TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC13TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC14TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC15TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC16TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC17TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaD4TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD5TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD6TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD7TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD8TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD9TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD10TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD11TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD12TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD13TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD14TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD15TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD16TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD17TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaE4TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE5TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE6TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE7TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE8TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE9TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE10TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE11TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE12TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE13TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE14TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE15TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE16TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE17TABLABALANCE = tabla_balance_df.iloc[8, 4]
    
    # Insertar el valor en la celda J6 (columna J, fila 6)
    ws["J6"] = 1
    
    # Condicion 2: G5 es "PPA"
    if g5_value == "PPA":
        celdaO19 = modelo_df.iloc[18, 14]
        celdaO42 = modelo_df.iloc[41, 14]
        celdaE3HR = hr_df.iloc[2, 4]
        celdaC14HR = hr_df.iloc[13, 2]
        celdaD8HR = hr_df.iloc[7, 3]
        celdaE8HR = hr_df.iloc[7, 4]
        celdaF8HR = hr_df.iloc[7, 5]
        celdaG8HR = hr_df.iloc[7, 6]
        celdaH8HR = hr_df.iloc[7, 7]
        celdaI8HR = hr_df.iloc[7, 8]
        celdaJ8HR = hr_df.iloc[7, 9]
        celdaK8HR = hr_df.iloc[7, 10]
        celdaL8HR = hr_df.iloc[7, 11]
        celdaM8HR = hr_df.iloc[7, 12]
        celdaN8HR = hr_df.iloc[7, 13]
        celdaO8HR = hr_df.iloc[7, 14]
        celdaP8HR = hr_df.iloc[7, 15]
        celdaQ8HR = hr_df.iloc[7, 16]
        celdaR8HR = hr_df.iloc[7, 17]
        celdaD17HR = hr_df.iloc[16, 3]
        celdaE17HR = hr_df.iloc[16, 4]
        celdaF17HR = hr_df.iloc[16, 5]
        celdaG17HR = hr_df.iloc[16, 6]
        celdaH17HR = hr_df.iloc[16, 7]
        celdaI17HR = hr_df.iloc[16, 8]
        celdaJ17HR = hr_df.iloc[16, 9]
        celdaK17HR = hr_df.iloc[16, 10]
        celdaL17HR = hr_df.iloc[16, 11]
        celdaM17HR = hr_df.iloc[16, 12]
        celdaN17HR = hr_df.iloc[16, 13]
        celdaO17HR = hr_df.iloc[16, 14]
        celdaP17HR = hr_df.iloc[16, 15]
        celdaQ17HR = hr_df.iloc[16, 16]
        celdaR17HR = hr_df.iloc[16, 17]
        
        # Insertar el valor en la celda J6 (columna J, fila 6)
        ws["J6"] = 1

    # Condicion 3: G5 es "VD"
    if g5_value == "VD":
        celdaK3 = modelo_df.iloc[2, 10]
        celdaK5 = modelo_df.iloc[4, 10]

        # Insertar el valor en la celda J6 (columna J, fila 6)
        ws["J6"] = 1





# Condicion 1: O1 es "2023-11EC"
if o1_value in ["2023-11EC"]:
    print("HOLA5")
    celdaE2 = modelo_df.iloc[1, 4]
    celdaG4 = modelo_df.iloc[3, 6]
    celdaG5 = modelo_df.iloc[4, 6]
    celdaG48 = modelo_df.iloc[47, 6]
    celdaG42 = modelo_df.iloc[41, 6]
    celdaG45 = modelo_df.iloc[44, 6]
    celdaG43 = modelo_df.iloc[42, 6]
    celdaG44 = modelo_df.iloc[43, 6]
    celdaG47 = modelo_df.iloc[46, 6]
    celdaG37 = modelo_df.iloc[36, 6]
    celdaG38 = modelo_df.iloc[37, 6]
    celdaG46 = modelo_df.iloc[45, 6]
    celdaO20 = modelo_df.iloc[19, 14]
    celdaO21 = modelo_df.iloc[20, 14]
    celdaB4TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB5TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB6TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB7TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB8TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB9TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB10TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB11TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB12TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB13TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB14TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB15TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB16TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB17TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaC4TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC5TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC6TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC7TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC8TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC9TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC10TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC11TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC12TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC13TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC14TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC15TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC16TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC17TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaD4TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD5TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD6TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD7TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD8TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD9TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD10TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD11TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD12TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD13TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD14TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD15TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD16TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD17TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaE4TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE5TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE6TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE7TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE8TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE9TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE10TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE11TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE12TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE13TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE14TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE15TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE16TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE17TABLABALANCE = tabla_balance_df.iloc[8, 4]
    
    # Insertar el valor en la celda J6 (columna J, fila 6)
    ws["J6"] = 1
        
    # Condicion 2: G5 es "PPA"
    if g5_value == "PPA":
        celdaO24 = modelo_df.iloc[23, 14]
        celdaO47 = modelo_df.iloc[46, 14]
        celdaE3HR = hr_df.iloc[2, 4]
        celdaC14HR = hr_df.iloc[13, 2]
        celdaD8HR = hr_df.iloc[7, 3]
        celdaE8HR = hr_df.iloc[7, 4]
        celdaF8HR = hr_df.iloc[7, 5]
        celdaG8HR = hr_df.iloc[7, 6]
        celdaH8HR = hr_df.iloc[7, 7]
        celdaI8HR = hr_df.iloc[7, 8]
        celdaJ8HR = hr_df.iloc[7, 9]
        celdaK8HR = hr_df.iloc[7, 10]
        celdaL8HR = hr_df.iloc[7, 11]
        celdaM8HR = hr_df.iloc[7, 12]
        celdaN8HR = hr_df.iloc[7, 13]
        celdaO8HR = hr_df.iloc[7, 14]
        celdaP8HR = hr_df.iloc[7, 15]
        celdaQ8HR = hr_df.iloc[7, 16]
        celdaR8HR = hr_df.iloc[7, 17]
        celdaD17HR = hr_df.iloc[16, 3]
        celdaE17HR = hr_df.iloc[16, 4]
        celdaF17HR = hr_df.iloc[16, 5]
        celdaG17HR = hr_df.iloc[16, 6]
        celdaH17HR = hr_df.iloc[16, 7]
        celdaI17HR = hr_df.iloc[16, 8]
        celdaJ17HR = hr_df.iloc[16, 9]
        celdaK17HR = hr_df.iloc[16, 10]
        celdaL17HR = hr_df.iloc[16, 11]
        celdaM17HR = hr_df.iloc[16, 12]
        celdaN17HR = hr_df.iloc[16, 13]
        celdaO17HR = hr_df.iloc[16, 14]
        celdaP17HR = hr_df.iloc[16, 15]
        celdaQ17HR = hr_df.iloc[16, 16]
        celdaR17HR = hr_df.iloc[16, 17]

        # Insertar el valor en la celda J6 (columna J, fila 6)
        ws["J6"] = 1

    # Condicion 3: G5 es "VD"
    if g5_value == "VD":
        celdaK3 = modelo_df.iloc[2, 10]
        celdaK5 = modelo_df.iloc[4, 10]
        
        # Insertar el valor en la celda J6 (columna J, fila 6)
        ws["J6"] = 1





# Condicion 1: O1 es "2022-45EC"
if o1_value in ["2022-45EC"]:
    print("HOLA6")
    celdaE2 = modelo_df.iloc[1, 4]
    celdaG4 = modelo_df.iloc[3, 6]
    celdaG5 = modelo_df.iloc[4, 6]
    celdaG47 = modelo_df.iloc[46, 6]
    celdaG41 = modelo_df.iloc[40, 6]
    celdaG44 = modelo_df.iloc[43, 6]
    celdaG42 = modelo_df.iloc[41, 6]
    celdaG34 = modelo_df.iloc[33, 6]
    celdaG46 = modelo_df.iloc[45, 6]
    celdaG36 = modelo_df.iloc[35, 6]
    celdaG37 = modelo_df.iloc[36, 6]
    celdaG45 = modelo_df.iloc[44, 6]
    celdaO20 = modelo_df.iloc[19, 14]
    celdaO21 = modelo_df.iloc[20, 14]
    celdaB4TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB5TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB6TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB7TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB8TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB9TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB10TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB11TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB12TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB13TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB14TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB15TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB16TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB17TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaC4TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC5TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC6TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC7TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC8TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC9TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC10TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC11TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC12TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC13TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC14TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC15TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC16TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC17TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaD4TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD5TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD6TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD7TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD8TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD9TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD10TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD11TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD12TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD13TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD14TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD15TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD16TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD17TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaE4TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE5TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE6TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE7TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE8TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE9TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE10TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE11TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE12TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE13TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE14TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE15TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE16TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE17TABLABALANCE = tabla_balance_df.iloc[8, 4]

    # Insertar el valor en la celda J6 (columna J, fila 6)
    ws["J6"] = 1    

    # Condicion 2: G5 es "PPA"
    if g5_value == "PPA":
        celdaO24 = modelo_df.iloc[23, 14]
        celdaO47 = modelo_df.iloc[46, 14]
        celdaE3HR = hr_df.iloc[2, 4]
        celdaC14HR = hr_df.iloc[13, 2]
        celdaD8HR = hr_df.iloc[7, 3]
        celdaE8HR = hr_df.iloc[7, 4]
        celdaF8HR = hr_df.iloc[7, 5]
        celdaG8HR = hr_df.iloc[7, 6]
        celdaH8HR = hr_df.iloc[7, 7]
        celdaI8HR = hr_df.iloc[7, 8]
        celdaJ8HR = hr_df.iloc[7, 9]
        celdaK8HR = hr_df.iloc[7, 10]
        celdaL8HR = hr_df.iloc[7, 11]
        celdaM8HR = hr_df.iloc[7, 12]
        celdaN8HR = hr_df.iloc[7, 13]
        celdaO8HR = hr_df.iloc[7, 14]
        celdaP8HR = hr_df.iloc[7, 15]
        celdaQ8HR = hr_df.iloc[7, 16]
        celdaR8HR = hr_df.iloc[7, 17]
        celdaD17HR = hr_df.iloc[16, 3]
        celdaE17HR = hr_df.iloc[16, 4]
        celdaF17HR = hr_df.iloc[16, 5]
        celdaG17HR = hr_df.iloc[16, 6]
        celdaH17HR = hr_df.iloc[16, 7]
        celdaI17HR = hr_df.iloc[16, 8]
        celdaJ17HR = hr_df.iloc[16, 9]
        celdaK17HR = hr_df.iloc[16, 10]
        celdaL17HR = hr_df.iloc[16, 11]
        celdaM17HR = hr_df.iloc[16, 12]
        celdaN17HR = hr_df.iloc[16, 13]
        celdaO17HR = hr_df.iloc[16, 14]
        celdaP17HR = hr_df.iloc[16, 15]
        celdaQ17HR = hr_df.iloc[16, 16]
        celdaR17HR = hr_df.iloc[16, 17]
        
        # Insertar el valor en la celda J6 (columna J, fila 6)
        ws["J6"] = 1

    # Condicion 3: G5 es "VD"
    if g5_value == "VD":
        celdaK3 = modelo_df.iloc[2, 10]
        celdaK4 = modelo_df.iloc[3, 10]
        
        # Insertar el valor en la celda J6 (columna J, fila 6)
        ws["J6"] = 1





# Condicion 1: O1 es "2022-451EC"
if o1_value in ["2022-41EC"]:
    print("HOLA7")
    celdaE2 = modelo_df.iloc[1, 4]
    celdaG4 = modelo_df.iloc[3, 6]
    celdaG5 = modelo_df.iloc[4, 6]
    celdaG47 = modelo_df.iloc[46, 6]
    celdaG41 = modelo_df.iloc[40, 6]
    celdaG44 = modelo_df.iloc[43, 6]
    celdaG42 = modelo_df.iloc[41, 6]
    celdaG33 = modelo_df.iloc[32, 6]
    celdaG46 = modelo_df.iloc[45, 6]
    celdaG36 = modelo_df.iloc[35, 6]
    celdaG37 = modelo_df.iloc[36, 6]
    celdaG45 = modelo_df.iloc[44, 6]
    celdaO20 = modelo_df.iloc[19, 14]
    celdaO21 = modelo_df.iloc[20, 14]
    celdaB4TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB5TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB6TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB7TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB8TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB9TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB10TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB11TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB12TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB13TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB14TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB15TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB16TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB17TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaC4TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC5TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC6TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC7TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC8TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC9TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC10TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC11TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC12TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC13TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC14TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC15TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC16TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC17TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaD4TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD5TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD6TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD7TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD8TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD9TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD10TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD11TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD12TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD13TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD14TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD15TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD16TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD17TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaE4TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE5TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE6TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE7TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE8TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE9TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE10TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE11TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE12TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE13TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE14TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE15TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE16TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE17TABLABALANCE = tabla_balance_df.iloc[8, 4]

    # Insertar el valor en la celda J6 (columna J, fila 6)
    ws["J6"] = 1
    
    # Condicion 2: G5 es "PPA"
    if g5_value == "PPA":
        celdaO24 = modelo_df.iloc[23, 14]
        celdaO47 = modelo_df.iloc[46, 14]
        celdaE3HR = hr_df.iloc[2, 4]
        celdaC14HR = hr_df.iloc[13, 2]
        celdaD8HR = hr_df.iloc[7, 3]
        celdaE8HR = hr_df.iloc[7, 4]
        celdaF8HR = hr_df.iloc[7, 5]
        celdaG8HR = hr_df.iloc[7, 6]
        celdaH8HR = hr_df.iloc[7, 7]
        celdaI8HR = hr_df.iloc[7, 8]
        celdaJ8HR = hr_df.iloc[7, 9]
        celdaK8HR = hr_df.iloc[7, 10]
        celdaL8HR = hr_df.iloc[7, 11]
        celdaM8HR = hr_df.iloc[7, 12]
        celdaN8HR = hr_df.iloc[7, 13]
        celdaO8HR = hr_df.iloc[7, 14]
        celdaP8HR = hr_df.iloc[7, 15]
        celdaQ8HR = hr_df.iloc[7, 16]
        celdaR8HR = hr_df.iloc[7, 17]
        celdaD17HR = hr_df.iloc[16, 3]
        celdaE17HR = hr_df.iloc[16, 4]
        celdaF17HR = hr_df.iloc[16, 5]
        celdaG17HR = hr_df.iloc[16, 6]
        celdaH17HR = hr_df.iloc[16, 7]
        celdaI17HR = hr_df.iloc[16, 8]
        celdaJ17HR = hr_df.iloc[16, 9]
        celdaK17HR = hr_df.iloc[16, 10]
        celdaL17HR = hr_df.iloc[16, 11]
        celdaM17HR = hr_df.iloc[16, 12]
        celdaN17HR = hr_df.iloc[16, 13]
        celdaO17HR = hr_df.iloc[16, 14]
        celdaP17HR = hr_df.iloc[16, 15]
        celdaQ17HR = hr_df.iloc[16, 16]
        celdaR17HR = hr_df.iloc[16, 17]

        # Insertar el valor en la celda J6 (columna J, fila 6)
        ws["J6"] = 1

    # Condicion 3: G5 es "VD"
    if g5_value == "VD":
        celdaK3 = modelo_df.iloc[2, 10]
        celdaK4 = modelo_df.iloc[3, 10]
        
        # Insertar el valor en la celda J6 (columna J, fila 6)
        ws["J6"] = 1





# Condicion 1: O1 es "2022-30"
if o1_value in ["2022-30"]:
    print("HOLA8")
    celdaE2 = modelo_df.iloc[1, 4]
    celdaG4 = modelo_df.iloc[3, 6]
    celdaG5 = modelo_df.iloc[4, 6]
    celdaG46 = modelo_df.iloc[45, 6]
    celdaG40 = modelo_df.iloc[39, 6]
    celdaG43 = modelo_df.iloc[42, 6]
    celdaG41 = modelo_df.iloc[40, 6]
    celdaG33 = modelo_df.iloc[32, 6]
    celdaG45 = modelo_df.iloc[44, 6]
    celdaG35 = modelo_df.iloc[34, 6]
    celdaG36 = modelo_df.iloc[35, 6]
    celdaG44 = modelo_df.iloc[43, 6]
    celdaO20 = modelo_df.iloc[19, 14]
    celdaO21 = modelo_df.iloc[20, 14]
    celdaB4TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB5TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB6TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB7TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB8TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB9TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB10TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB11TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB12TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB13TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB14TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB15TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB16TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaB17TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaC4TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC5TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC6TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC7TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC8TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC9TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC10TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC11TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC12TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC13TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC14TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC15TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC16TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaC17TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaD4TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD5TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD6TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD7TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD8TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD9TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD10TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD11TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD12TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD13TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD14TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD15TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD16TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaD17TABLABALANCE = tabla_balance_df.iloc[5, 4]
    celdaE4TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE5TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE6TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE7TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE8TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE9TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE10TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE11TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE12TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE13TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE14TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE15TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE16TABLABALANCE = tabla_balance_df.iloc[8, 4]
    celdaE17TABLABALANCE = tabla_balance_df.iloc[8, 4]

    # Insertar el valor en la celda J6 (columna J, fila 6)
    ws["J6"] = 1
    
    # Condicion 2: G5 es "PPA"
    if g5_value == "PPA":
        celdaO24 = modelo_df.iloc[23, 14]
        celdaO47 = modelo_df.iloc[46, 14]
        celdaE3HR = hr_df.iloc[2, 4]
        celdaC14HR = hr_df.iloc[13, 2]
        celdaD8HR = hr_df.iloc[7, 3]
        celdaE8HR = hr_df.iloc[7, 4]
        celdaF8HR = hr_df.iloc[7, 5]
        celdaG8HR = hr_df.iloc[7, 6]
        celdaH8HR = hr_df.iloc[7, 7]
        celdaI8HR = hr_df.iloc[7, 8]
        celdaJ8HR = hr_df.iloc[7, 9]
        celdaK8HR = hr_df.iloc[7, 10]
        celdaL8HR = hr_df.iloc[7, 11]
        celdaM8HR = hr_df.iloc[7, 12]
        celdaN8HR = hr_df.iloc[7, 13]
        celdaO8HR = hr_df.iloc[7, 14]
        celdaP8HR = hr_df.iloc[7, 15]
        celdaQ8HR = hr_df.iloc[7, 16]
        celdaR8HR = hr_df.iloc[7, 17]
        celdaD17HR = hr_df.iloc[16, 3]
        celdaE17HR = hr_df.iloc[16, 4]
        celdaF17HR = hr_df.iloc[16, 5]
        celdaG17HR = hr_df.iloc[16, 6]
        celdaH17HR = hr_df.iloc[16, 7]
        celdaI17HR = hr_df.iloc[16, 8]
        celdaJ17HR = hr_df.iloc[16, 9]
        celdaK17HR = hr_df.iloc[16, 10]
        celdaL17HR = hr_df.iloc[16, 11]
        celdaM17HR = hr_df.iloc[16, 12]
        celdaN17HR = hr_df.iloc[16, 13]
        celdaO17HR = hr_df.iloc[16, 14]
        celdaP17HR = hr_df.iloc[16, 15]
        celdaQ17HR = hr_df.iloc[16, 16]
        celdaR17HR = hr_df.iloc[16, 17]

        # Insertar el valor en la celda J6 (columna J, fila 6)
        ws["J6"] = 1

    # Condicion 3: G5 es "VD"
    if g5_value == "VD":
        celdaK3 = modelo_df.iloc[2, 10]
        celdaK4 = modelo_df.iloc[3, 10]

        # Insertar el valor en la celda J6 (columna J, fila 6)
        ws["J6"] = 1





# Guardar los cambios en el archivo
wb.save(file_path2)